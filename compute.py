#!/usr/bin/env python3
"""Compute dashboard data for GitHub Actions.
   Downloads ARR working spreadsheet from SharePoint.
   Extracts BCL (big customer list), CEO dashboard, UP explorer data.
   Writes to dashboard_data.json.
"""

import json
import os
import datetime
import calendar
from collections import defaultdict
from io import BytesIO
from openpyxl import load_workbook
import requests


# SharePoint URL (new working file)
SOURCE_URL = (
    "https://wiseandsallycom-my.sharepoint.com/:x:/g/personal/"
    "csongor_doma_cambri_io/"
    "IQCtH2LmrI_CTL0I5sd7-TY-AZtRdW7nZn-vjGaH1tBhEf4"
    "?e=EwShbU&download=1"
)


def download_workbook(url, name):
    """Download Excel workbook from SharePoint URL."""
    print(f"  Downloading {name}...")
    session = requests.Session()
    session.headers["User-Agent"] = (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0 Safari/537.36"
    )
    try:
        resp = session.get(url, allow_redirects=True, timeout=120)
        resp.raise_for_status()
        print(f"    Downloaded {len(resp.content)} bytes")
        return load_workbook(BytesIO(resp.content), data_only=True, read_only=True)
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Failed to download {name}: {e}")


def safe_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def safe_str(v):
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s in ('#VALUE!', '#N/A') else s


def safe_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    return ''


def fmt_date_display(d):
    if isinstance(d, datetime.datetime):
        return d.strftime('%d/%m/%Y')
    if isinstance(d, datetime.date):
        return d.strftime('%d/%m/%Y')
    return str(d) if d else ''


# ═══════════════════════════════════════════════════════════════
#  ACCOUNTS
# ═══════════════════════════════════════════════════════════════

CUSTOMER_TYPES = {
    'Customer', 'Customer - Dormant (90days)', 'Customer - Pilot',
    'Customer - Repeat', 'License Client'
}


def read_accounts(wb):
    """Read accounts from 'Accs for ARR Work' sheet.

    New columns (row 5 = headers, data from row 6):
      Col 4  (E)  Account Owner
      Col 5  (F)  Account Name
      Col 6  (G)  TAM Type
      Col 7  (H)  Type
      Col 8  (I)  TAM Account
      Col 11 (L)  Billing Country
      Col 12 (M)  Industry c
      Col 16 (Q)  Acc Casesafe ID 18
      Col 18 (S)  Ultimate Parent ID casesafe 18
      Col 19 (T)  Customer Success Manager
      Col 27 (AB) Hierarchy Live Total ARR (converted)
      Col 29 (AD) Hierarchy Live B1 ARR (converted)
      Col 31 (AF) Hierarchy Live B2 ARR (converted)
      Col 33 (AH) Live Total ARR (converted)
      Col 35 (AJ) Live B1 ARR (converted)
      Col 37 (AL) Live B2 ARR (converted)
      Col 39 (AN) Current Credit Balance Total (converted)
      Col 44 (AS) Parent Account
      Col 45 (AT) Parent Account ID
      Col 46 (AU) Account Cohort
    """
    ws = wb['Accs for ARR Work']
    accounts = {}
    id15_to_id18 = {}
    acc_casesafe_to_up = {}

    for _row in ws.iter_rows(min_row=6, values_only=True):
        row = list(_row)
        if len(row) < 19:
            continue
        acc_id = str(row[16] or '').strip()   # Q = Acc Casesafe ID 18
        if not acc_id:
            continue
        name = safe_str(row[5])               # F = Account Name
        acc_type = safe_str(row[7])           # H = Type
        up_id = str(row[18] or '').strip()    # S = UP ID casesafe 18

        accounts[acc_id] = {
            'name': name,
            'type': acc_type,
            'country': safe_str(row[11]),     # L = Billing Country
            'parent_id': safe_str(row[45]) if len(row) > 45 else '',  # AT = Parent Account ID
            'up_id': up_id,
            'owner': safe_str(row[4]),        # E = Account Owner
            'csm': safe_str(row[19]) if len(row) > 19 else '',  # T = CSM
            'industry': safe_str(row[12]),    # M = Industry c
            'employees': None,                # not in new file
            'last_activity': '',              # not in new file
            'credit_balance': safe_float(row[39]) if len(row) > 39 else 0,  # AN
            'rev_target_2026': 0,             # not in new file
            'is_customer': acc_type in CUSTOMER_TYPES,
            'b1_arr': safe_float(row[35]) if len(row) > 35 else 0,  # AJ = Live B1
            'b2_arr': safe_float(row[37]) if len(row) > 37 else 0,  # AL = Live B2
            'total_arr': safe_float(row[33]) if len(row) > 33 else 0,  # AH = Live Total
            'hierarchy_total_arr': safe_float(row[27]) if len(row) > 27 else 0,  # AB
            'hierarchy_b1_arr': safe_float(row[29]) if len(row) > 29 else 0,  # AD
            'hierarchy_b2_arr': safe_float(row[31]) if len(row) > 31 else 0,  # AF
            'eoq4_total': safe_float(row[21]) if len(row) > 21 else 0,  # V = EoQ4 2025 ARR Total
            'eoq4_b1': safe_float(row[23]) if len(row) > 23 else 0,    # X = EoQ4 2025 ARR B1
            'eoq4_b2': safe_float(row[25]) if len(row) > 25 else 0,    # Z = EoQ4 2025 ARR B2
            'tam_type': safe_str(row[6]),     # G = TAM Type
            'account_cohort': safe_str(row[46]) if len(row) > 46 else '',  # AU = Account Cohort
        }
        acc_casesafe_to_up[acc_id] = up_id
        if len(acc_id) >= 15:
            id15_to_id18[acc_id[:15]] = acc_id

    print(f"    {len(accounts)} accounts")
    return accounts, id15_to_id18, acc_casesafe_to_up


# ═══════════════════════════════════════════════════════════════
#  CONTRACT ITEMS
# ═══════════════════════════════════════════════════════════════

def read_contract_items(wb, acc_casesafe_to_up):
    """Read CIs from 'CI report for ARR' sheet.

    New columns (row 5 = headers, data from row 6):
      Col 4  (E)  Contract Items Name
      Col 5  (F)  Description
      Col 6  (G)  Product Family
      Col 8  (I)  Billed Value (converted)
      Col 9  (J)  Invoice Date
      Col 10 (K)  Start Date
      Col 11 (L)  End Date
      Col 12 (M)  Billing Status
      Col 13 (N)  Active Contract Item
      Col 14 (O)  ARR Bucket  (B1, B2, or empty)
      Col 16 (Q)  ARR B2 (converted)
      Col 18 (S)  ARR B1 (converted)
      Col 19 (T)  Account
      Col 27 (AB) Contract Line Casesafe 18
      Col 28 (AC) Account Casesafe ID 18
      Col 29 (AD) UP Name
      Col 30 (AE) UP ID
    """
    ws = wb['CI report for ARR']
    all_ci = []
    BUCKET_ORDER = {'B1': 0, 'B2': 1, 'Excluded': 2, '': 3, '—': 3}

    EXCLUDED_BILLING = {'Pending Details', 'Data Loaded Back Data', 'Not to be Invoiced'}

    for _row in ws.iter_rows(min_row=6, values_only=True):
        row = list(_row)
        if len(row) < 28:
            continue

        ci_name = safe_str(row[4])
        if not ci_name:
            continue

        pf = safe_str(row[6])
        bv = safe_float(row[8])
        start_raw = row[10]
        end_raw = row[11]
        billing_status = safe_str(row[12])
        active = row[13]
        bucket_raw = safe_str(row[14])  # Pre-computed: B1, B2, or empty
        arr_b2 = safe_float(row[16])
        arr_b1 = safe_float(row[18])
        account = safe_str(row[19])
        credit_balance = safe_float(row[22]) if len(row) > 22 else 0  # Col 22 (W) = Credit Balance (converted)
        ci_casesafe18 = safe_str(row[27])
        acc_id18 = safe_str(row[28])
        up_name = safe_str(row[29])
        up_id = safe_str(row[30])

        # Determine bucket
        if bucket_raw in ('B1', 'B2'):
            bucket = bucket_raw
        elif billing_status in EXCLUDED_BILLING:
            bucket = 'Excluded'
        else:
            bucket = '—'

        # Determine contributing status
        contributing = 1 if bucket in ('B1', 'B2') else 0

        # If UP ID not in this row, try to get it from accounts
        if not up_id and acc_id18:
            up_id = acc_casesafe_to_up.get(acc_id18, '')

        all_ci.append({
            'up_id': up_id,
            'up_name': up_name,
            'bucket': bucket,
            'account': account,
            'account_id': acc_id18,
            'name': ci_name,
            'ci_id': ci_casesafe18,
            'product_family': pf,
            'billed_value': bv,
            'arr_b1': arr_b1,
            'arr_b2': arr_b2,
            'billing_status': billing_status,
            'start_date': fmt_date_display(start_raw),
            'end_date': fmt_date_display(end_raw),
            'start_iso': safe_date(start_raw),
            'end_iso': safe_date(end_raw),
            'active': 1 if active else 0,
            'contributing': contributing,
            'credit_balance': credit_balance,
        })

    print(f"    {len(all_ci)} contract items")
    return all_ci


# ═══════════════════════════════════════════════════════════════
#  SAMPLES (testing revenue + CEO dashboard)
# ═══════════════════════════════════════════════════════════════

def read_samples(wb, acc_casesafe_to_up, accounts):
    """Read samples from 'All Samples All Info 2' sheet.

    New columns (row 5 = headers, data from row 6):
      Col 5  (F)  Account
      Col 8  (I)  Completed Sample Size
      Col 10 (K)  Revenue (converted)
      Col 12 (M)  Status
      Col 14 (O)  Date Completed
      Col 18 (S)  Account Casesafe ID 18
    """
    ws = wb['All Samples All Info 2']

    EXCLUDED_STATUSES_CEO = {'Not reconciled', 'Not to be Invoiced', '', 'None'}
    EXCLUDED_STATUSES_TESTING = {'Not to be Invoiced', 'Not reconciled', 'Data Loaded Back Data'}
    EXCLUDED_STATUSES_YTD = {'Not to be Invoiced'}

    today = datetime.date.today()
    current_year = today.year
    current_month = today.month

    # YTD testing revenue (status != "Not to be Invoiced", date_completed up to today's day-of-year)
    ytd_testing_this_year = 0.0
    ytd_testing_last_year = 0.0
    # Per-account YTD testing revenue for grouping
    acc_ytd_ty = defaultdict(float)  # acc_id18 -> TYTD
    acc_ytd_ly = defaultdict(float)  # acc_id18 -> LYTD

    # Active customers with same-day cutoff (CEO exclusion criteria)
    active_ups_ytd = set()
    active_ups_lytd = set()
    # The cutoff date for last year is the same month/day but in the previous year
    try:
        ly_cutoff = datetime.date(current_year - 1, today.month, today.day)
    except ValueError:
        # Handle Feb 29 edge case
        ly_cutoff = datetime.date(current_year - 1, today.month, 28)

    # CEO aggregation
    company_monthly_rev = defaultdict(float)
    company_daily_rev = defaultdict(lambda: defaultdict(float))

    # Testing revenue by UP
    up_monthly_rev = defaultdict(lambda: defaultdict(float))
    up_yearly_rev = defaultdict(lambda: defaultdict(float))
    up_total_rev = defaultdict(float)

    # Testing revenue by account
    acc_monthly_rev = defaultdict(lambda: defaultdict(float))
    acc_yearly_rev = defaultdict(lambda: defaultdict(float))
    acc_total_rev = defaultdict(float)

    # UP yearly for active customers count
    up_yearly_rev_ceo = defaultdict(lambda: defaultdict(float))

    for _row in ws.iter_rows(min_row=6, values_only=True):
        row = list(_row)
        if len(row) < 15:
            continue

        acc_name = safe_str(row[5])     # F = Account
        rev = row[10]                    # K = Revenue (converted)
        status = safe_str(row[12])       # M = Status
        date_completed = row[14]         # O = Date Completed
        acc_id18 = safe_str(row[18]) if len(row) > 18 else ''  # S = Account Casesafe ID 18

        if not date_completed or not rev:
            continue

        try:
            rev_f = float(rev)
        except (ValueError, TypeError):
            continue

        # Parse date
        if isinstance(date_completed, datetime.datetime):
            yr, mo, dy = date_completed.year, date_completed.month, date_completed.day
        elif isinstance(date_completed, datetime.date):
            yr, mo, dy = date_completed.year, date_completed.month, date_completed.day
        else:
            continue

        if yr > 2100 or yr < 2000:
            continue

        key = f"{yr}-{mo:02d}"

        # Resolve UP name via account ID -> accounts -> UP ID -> accounts[UP ID].name
        up_id = acc_casesafe_to_up.get(acc_id18, '')
        up_name = ''
        if up_id and up_id in accounts:
            up_name = accounts[up_id].get('name', '')
        elif acc_name:
            # Fallback: try to find account by name
            for aid, ainfo in accounts.items():
                if ainfo['name'] == acc_name:
                    up_id = ainfo.get('up_id', '')
                    if up_id and up_id in accounts:
                        up_name = accounts[up_id].get('name', '')
                    break

        # CEO dashboard aggregation (exclude Not reconciled, Not to be Invoiced, blank)
        if status not in EXCLUDED_STATUSES_CEO:
            company_monthly_rev[key] += rev_f
            company_daily_rev[key][dy] += rev_f
            if up_name:
                up_yearly_rev_ceo[up_name][yr] += rev_f
                # Track active customers with same-day cutoff
                sample_date_ceo = datetime.date(yr, mo, dy)
                if yr == current_year and sample_date_ceo <= today:
                    active_ups_ytd.add(up_name)
                elif yr == current_year - 1 and sample_date_ceo <= ly_cutoff:
                    active_ups_lytd.add(up_name)

        # Testing revenue aggregation (exclude Not to be Invoiced, Not reconciled, Data Loaded Back Data)
        if status not in EXCLUDED_STATUSES_TESTING:
            if up_name:
                up_monthly_rev[up_name][key] += rev_f
                up_yearly_rev[up_name][yr] += rev_f
                up_total_rev[up_name] += rev_f
            if acc_name:
                acc_monthly_rev[acc_name][key] += rev_f
                acc_yearly_rev[acc_name][yr] += rev_f
                acc_total_rev[acc_name] += rev_f

        # YTD testing revenue (status != "Not to be Invoiced")
        if status not in EXCLUDED_STATUSES_YTD:
            sample_date = datetime.date(yr, mo, dy)
            if yr == current_year and sample_date <= today:
                ytd_testing_this_year += rev_f
                if acc_id18:
                    acc_ytd_ty[acc_id18] += rev_f
            elif yr == current_year - 1 and sample_date <= ly_cutoff:
                ytd_testing_last_year += rev_f
                if acc_id18:
                    acc_ytd_ly[acc_id18] += rev_f

    print(f"    {len(up_total_rev)} UPs with testing revenue")
    print(f"    {len(company_monthly_rev)} months of CEO data")

    return {
        'company_monthly_rev': company_monthly_rev,
        'company_daily_rev': company_daily_rev,
        'up_monthly_rev': up_monthly_rev,
        'up_yearly_rev': up_yearly_rev,
        'up_total_rev': up_total_rev,
        'acc_monthly_rev': acc_monthly_rev,
        'acc_yearly_rev': acc_yearly_rev,
        'acc_total_rev': acc_total_rev,
        'up_yearly_rev_ceo': up_yearly_rev_ceo,
        'ytd_testing_this_year': ytd_testing_this_year,
        'ytd_testing_last_year': ytd_testing_last_year,
        'acc_ytd_ty': dict(acc_ytd_ty),
        'acc_ytd_ly': dict(acc_ytd_ly),
        'active_ups_ytd': len(active_ups_ytd),
        'active_ups_lytd': len(active_ups_lytd),
    }


# ═══════════════════════════════════════════════════════════════
#  CEO DASHBOARD
# ═══════════════════════════════════════════════════════════════

def extract_ceo_dashboard(samples_data, accounts, acc_casesafe_to_up):
    """Build CEO dashboard from sample aggregations."""
    print("  Building CEO dashboard...")
    today = datetime.date.today()
    current_year = today.year
    current_month = today.month

    company_monthly_rev = samples_data['company_monthly_rev']
    company_daily_rev = samples_data['company_daily_rev']
    up_yearly_rev_ceo = samples_data['up_yearly_rev_ceo']

    # Build CEO months (last 25 months)
    ceo_months = []
    for i in range(24, -1, -1):
        m = current_month - i
        y = current_year
        while m <= 0:
            m += 12
            y -= 1
        ceo_months.append(f"{y}-{m:02d}")

    ceo_monthly = {m: round(company_monthly_rev.get(m, 0), 2) for m in ceo_months}

    # Daily cumulative data (last 12 months)
    ceo_daily_cumulative = {}
    for i in range(11, -1, -1):
        m = current_month - i
        y = current_year
        while m <= 0:
            m += 12
            y -= 1
        key = f"{y}-{m:02d}"
        days_in_month = calendar.monthrange(y, m)[1]
        daily = company_daily_rev.get(key, {})
        # For current month, only include data up to today
        max_day = today.day if (y == current_year and m == current_month) else days_in_month
        cumulative = []
        running = 0.0
        for d in range(1, days_in_month + 1):
            running += daily.get(d, 0)
            if d <= max_day:
                cumulative.append(round(running, 2))
            # skip future days for current month
        ceo_daily_cumulative[key] = cumulative

    # KPIs
    lytd_total = sum(
        company_monthly_rev.get(f"{current_year - 1}-{m:02d}", 0)
        for m in range(1, current_month + 1)
    )
    tytd_total = sum(
        company_monthly_rev.get(f"{current_year}-{m:02d}", 0)
        for m in range(1, current_month + 1)
    )
    ytd_growth = (
        round((tytd_total / lytd_total - 1) * 100, 1) if lytd_total > 0 else None
    )

    # Active customers YTD (with same-day cutoff from read_samples)
    active_customers_ytd = samples_data.get('active_ups_ytd', 0)
    active_customers_lytd = samples_data.get('active_ups_lytd', 0)

    # Cumulative chart — build from daily data (replaces old formulas sheet)
    # We build series for the last 13 months (current + 12 prior)
    cum_series = {}
    for i in range(12, -1, -1):
        m = current_month - i
        y = current_year
        while m <= 0:
            m += 12
            y -= 1
        key = f"{y}-{m:02d}"
        label = datetime.date(y, m, 1).strftime('%b %y')
        if i == 0:
            label = 'This Month'
        elif i == 1:
            label = 'Last Month'
        cum_series[label] = ceo_daily_cumulative.get(key, [])

    # Find last actual data day for this month
    # For the current month, use today's day (data beyond today is just carry-forward)
    this_month_data = cum_series.get('This Month', [])
    cum_last_day = min(today.day, len(this_month_data)) if this_month_data else 0

    cum_this_mtd = this_month_data[cum_last_day - 1] if cum_last_day > 0 else None

    cum_chart = {
        'subtitle': f'{datetime.date(current_year, current_month, 1).strftime("%B %Y")} cumulative',
        'series': cum_series,
        'l8m_avg': [],
        'forecast': [],
        'last_day': cum_last_day,
        'this_mtd': cum_this_mtd,
        'avg_at_day': None,
    }

    # YTD testing revenue
    ytd_testing_ty = round(samples_data.get('ytd_testing_this_year', 0), 2)
    ytd_testing_ly = round(samples_data.get('ytd_testing_last_year', 0), 2)
    ytd_testing_growth = (
        round((ytd_testing_ty / ytd_testing_ly - 1) * 100, 1)
        if ytd_testing_ly > 0 else None
    )

    # ── Total ARR from UP-level accounts ──
    # Sum hierarchy ARR at UP level only (accounts where acc_id == up_id) to avoid double counting
    total_live_arr = 0.0
    total_eoq4_arr = 0.0
    seen_ups = set()
    for acc_id, info in accounts.items():
        up_id = acc_casesafe_to_up.get(acc_id, '')
        # Only count UP-level accounts (where the account IS the ultimate parent)
        if up_id and up_id == acc_id and up_id not in seen_ups:
            seen_ups.add(up_id)
            total_live_arr += info.get('hierarchy_total_arr', 0) or 0
            total_eoq4_arr += info.get('eoq4_total', 0) or 0

    arr_change_pct = (
        round((total_live_arr / total_eoq4_arr - 1) * 100, 1)
        if total_eoq4_arr > 0 else None
    )

    # ── Grouped testing revenue by dimension ──
    # Group per-account YTD revenue by CSM, Country, Industry, and Cohort
    acc_ytd_ty = samples_data.get('acc_ytd_ty', {})
    acc_ytd_ly = samples_data.get('acc_ytd_ly', {})

    # Collect all account IDs that have any YTD revenue
    all_acc_ids = set(acc_ytd_ty.keys()) | set(acc_ytd_ly.keys())

    # Build dimension lookups from accounts (resolve to UP-level where possible)
    dim_groups = {
        'cohort': defaultdict(lambda: {'tytd': 0.0, 'lytd': 0.0}),
        'csm': defaultdict(lambda: {'tytd': 0.0, 'lytd': 0.0}),
        'country': defaultdict(lambda: {'tytd': 0.0, 'lytd': 0.0}),
        'industry': defaultdict(lambda: {'tytd': 0.0, 'lytd': 0.0}),
        'tenure': defaultdict(lambda: {'tytd': 0.0, 'lytd': 0.0}),
    }

    # Pre-compute tenure per UP from up_monthly_rev (earliest month with revenue)
    up_monthly_rev = samples_data['up_monthly_rev']
    twelve_months_ago = f"{current_year - 1}-{today.month:02d}"
    up_tenure_cache = {}
    for up_name, mdata in up_monthly_rev.items():
        months_with_data = sorted(k for k, v in mdata.items() if v > 0)
        if months_with_data and months_with_data[0] < twelve_months_ago:
            up_tenure_cache[up_name] = 'Established Business'
        else:
            up_tenure_cache[up_name] = 'New Logo'

    for acc_id in all_acc_ids:
        ty = acc_ytd_ty.get(acc_id, 0)
        ly = acc_ytd_ly.get(acc_id, 0)
        if ty == 0 and ly == 0:
            continue

        # Get account info
        acc_info = accounts.get(acc_id, {})
        csm = acc_info.get('csm', '') or 'Unassigned'
        country = acc_info.get('country', '') or 'Unknown'
        industry = acc_info.get('industry', '') or 'Unknown'

        # Account Cohort from the sheet (column AU)
        cohort = acc_info.get('account_cohort', '') or 'Unknown'

        # Tenure from UP-level first test date
        up_id = acc_casesafe_to_up.get(acc_id, acc_id)
        up_name_for_tenure = accounts.get(up_id, {}).get('name', '')
        tenure_val = up_tenure_cache.get(up_name_for_tenure, 'New Logo')

        dim_groups['cohort'][cohort]['tytd'] += ty
        dim_groups['cohort'][cohort]['lytd'] += ly
        dim_groups['csm'][csm]['tytd'] += ty
        dim_groups['csm'][csm]['lytd'] += ly
        dim_groups['country'][country]['tytd'] += ty
        dim_groups['country'][country]['lytd'] += ly
        dim_groups['industry'][industry]['tytd'] += ty
        dim_groups['industry'][industry]['lytd'] += ly
        dim_groups['tenure'][tenure_val]['tytd'] += ty
        dim_groups['tenure'][tenure_val]['lytd'] += ly

    # Convert to serialisable format: list of {label, tytd, lytd, growth}
    grouped_testing_rev = {}
    for dim_key, groups in dim_groups.items():
        entries = []
        for label, vals in groups.items():
            tytd = round(vals['tytd'], 2)
            lytd = round(vals['lytd'], 2)
            growth = round((tytd / lytd - 1) * 100, 1) if lytd > 0 else (100.0 if tytd > 0 else 0.0)
            entries.append({'label': label, 'tytd': tytd, 'lytd': lytd, 'growth': growth})
        # Sort by TYTD descending
        entries.sort(key=lambda e: e['tytd'], reverse=True)
        grouped_testing_rev[dim_key] = entries

    print(
        f"    LYTD: {round(lytd_total):,}, TYTD: {round(tytd_total):,}, "
        f"Growth: {ytd_growth}%"
    )
    print(
        f"    Testing Revenue YTD: {round(ytd_testing_ty):,}, "
        f"LYTD: {round(ytd_testing_ly):,}, Growth: {ytd_testing_growth}%"
    )
    print(f"    Active customers YTD: {active_customers_ytd}, LYTD: {active_customers_lytd}")
    print(f"    Total Live ARR: {round(total_live_arr):,}, EoQ4 ARR: {round(total_eoq4_arr):,}, Change: {arr_change_pct}%")
    for dim, entries in grouped_testing_rev.items():
        print(f"    Grouped by {dim}: {len(entries)} groups")

    return {
        'months': ceo_months,
        'monthly_rev': ceo_monthly,
        'daily_cumulative': ceo_daily_cumulative,
        'cum_chart': cum_chart,
        'lytd': round(lytd_total, 2),
        'tytd': round(tytd_total, 2),
        'ytd_growth': ytd_growth,
        'active_customers_ytd': active_customers_ytd,
        'active_customers_lytd': active_customers_lytd,
        'ytd_testing_this_year': ytd_testing_ty,
        'ytd_testing_last_year': ytd_testing_ly,
        'ytd_testing_growth': ytd_testing_growth,
        'grouped_testing_rev': grouped_testing_rev,
        'total_live_arr': round(total_live_arr, 2),
        'total_eoq4_arr': round(total_eoq4_arr, 2),
        'arr_change_pct': arr_change_pct,
    }


# ═══════════════════════════════════════════════════════════════
#  UP EXPLORER (Customer Analysis tab)
# ═══════════════════════════════════════════════════════════════

def extract_up_explorer(accounts, id15_to_id18, acc_casesafe_to_up, all_ci, samples_data):
    """Build UP Explorer data from accounts, CIs, and samples."""
    print("  Building UP Explorer...")
    today = datetime.date.today()
    current_year = today.year
    current_month = today.month

    up_monthly_rev = samples_data['up_monthly_rev']
    up_yearly_rev = samples_data['up_yearly_rev']
    up_total_rev = samples_data['up_total_rev']
    acc_monthly_rev = samples_data['acc_monthly_rev']
    acc_yearly_rev = samples_data['acc_yearly_rev']
    acc_total_rev = samples_data['acc_total_rev']

    BUCKET_ORDER = {'B1': 0, 'B2': 1, 'Excluded': 2, '—': 3}

    # Group by UP
    up_groups = defaultdict(list)
    for acc_id, info in accounts.items():
        if info['up_id']:
            up_groups[info['up_id']].append(acc_id)

    # Build children map
    children_map = defaultdict(list)
    for acc_id, info in accounts.items():
        pid = info['parent_id']
        if pid:
            pid_str = str(pid).strip()
            pid18 = id15_to_id18.get(pid_str[:15] if pid_str else None)
            if pid18 and pid18 in accounts:
                children_map[pid18].append(acc_id)

    # DFS hierarchy builder
    def build_tree(up_id, acc_ids):
        acc_set = set(acc_ids)
        roots = []
        for aid in acc_ids:
            pid = accounts[aid]['parent_id']
            if not pid:
                roots.append(aid)
                continue
            pid_str = str(pid).strip()
            pid18 = id15_to_id18.get(pid_str[:15] if pid_str else None)
            if not pid18 or pid18 not in acc_set:
                roots.append(aid)
        roots.sort(key=lambda x: (accounts[x]['name'] or '').lower())

        def visit(node, level):
            info = accounts[node]
            result = [{
                'id': node, 'name': info['name'], 'type': info['type'],
                'country': info['country'], 'level': level,
                'owner': info['owner'], 'csm': info['csm'],
                'industry': info['industry'], 'is_customer': info['is_customer'],
                'last_activity': info['last_activity'],
                'credit_balance': info.get('credit_balance', 0) or 0,
                'rev_target_2026': info.get('rev_target_2026', 0) or 0,
                'b1_arr': info['b1_arr'], 'b2_arr': info['b2_arr'],
                'total_arr': info['total_arr'],
            }]
            kids = [c for c in children_map.get(node, []) if c in acc_set]
            kids.sort(key=lambda x: (accounts[x]['name'] or '').lower())
            for kid in kids:
                result.extend(visit(kid, level + 1))
            return result

        tree = []
        for root in roots:
            tree.extend(visit(root, 0))
        return tree

    # Group CIs by UP
    ci_by_up = defaultdict(list)
    for ci in all_ci:
        if ci['up_id']:
            ci_by_up[ci['up_id']].append(ci)
    for up_id in ci_by_up:
        ci_by_up[up_id].sort(key=lambda x: BUCKET_ORDER.get(x['bucket'], 99))

    # UP-level ARR from CIs
    up_b1_arr = defaultdict(float)
    up_b2_arr = defaultdict(float)
    up_b1_items = defaultdict(int)
    up_b2_items = defaultdict(int)
    for ci in all_ci:
        if ci['bucket'] == 'B1':
            up_b1_arr[ci['up_id']] += ci['arr_b1']
            up_b1_items[ci['up_id']] += 1
        elif ci['bucket'] == 'B2':
            up_b2_arr[ci['up_id']] += ci['arr_b2']
            up_b2_items[ci['up_id']] += 1

    up_data = []
    seen_ups = set()
    for up_id, items in ci_by_up.items():
        if not up_id or up_id in seen_ups:
            continue
        seen_ups.add(up_id)
        b1 = up_b1_arr.get(up_id, 0)
        b2 = up_b2_arr.get(up_id, 0)
        total = b1 + b2

        up_name = accounts.get(up_id, {}).get('name', '')
        if not up_name:
            up_name = items[0].get('up_name', up_id) if items else up_id
        acc_ids = up_groups.get(up_id, [])
        hierarchy = build_tree(up_id, acc_ids)
        up_accounts = sorted(set(ci['account'] for ci in items if ci['account']))
        up_billing_statuses = sorted(set(ci['billing_status'] for ci in items if ci['billing_status']))

        # Use Hierarchy Live ARR from the UP's own account record (top-level lookup)
        up_acc = accounts.get(up_id, {})
        h_b1 = up_acc.get('hierarchy_b1_arr', 0) or 0
        h_b2 = up_acc.get('hierarchy_b2_arr', 0) or 0
        h_total = up_acc.get('hierarchy_total_arr', 0) or 0
        eoq4_b1 = up_acc.get('eoq4_b1', 0) or 0
        eoq4_b2 = up_acc.get('eoq4_b2', 0) or 0
        eoq4_total = up_acc.get('eoq4_total', 0) or 0

        up_data.append({
            'id': up_id,
            'name': up_name,
            'b1_arr': round(h_b1, 2), 'b2_arr': round(h_b2, 2), 'total_arr': round(h_total, 2),
            'eoq4_b1': round(eoq4_b1, 2), 'eoq4_b2': round(eoq4_b2, 2), 'eoq4_total': round(eoq4_total, 2),
            'b1_items': up_b1_items.get(up_id, 0),
            'b2_items': up_b2_items.get(up_id, 0),
            'account_count': len(acc_ids),
            'ci_count': len(items),
            'hierarchy': hierarchy,
            'contract_items': items,
            'ci_accounts': up_accounts,
            'ci_billing_statuses': up_billing_statuses,
        })

    up_data.sort(key=lambda x: -x['total_arr'])
    print(f"    {len(up_data)} UPs with contract items")

    # Sparkline months (last 24)
    sparkline_months = []
    for i in range(23, -1, -1):
        m = current_month - i
        y = current_year
        while m <= 0:
            m += 12
            y -= 1
        sparkline_months.append(f"{y}-{m:02d}")

    def compute_ytd_metrics(monthly_dict):
        lytd = sum(monthly_dict.get(f"{current_year - 1}-{mo:02d}", 0) for mo in range(1, current_month + 1))
        tytd = sum(monthly_dict.get(f"{current_year}-{mo:02d}", 0) for mo in range(1, current_month + 1))
        nrr = round((tytd / lytd - 1) * 100) if lytd > 0 else None
        return round(lytd, 2), round(tytd, 2), nrr

    # Enrich UP data with testing revenue
    up_name_to_idx = {u['name']: i for i, u in enumerate(up_data)}
    matched = 0
    for up_name, total in up_total_rev.items():
        idx = up_name_to_idx.get(up_name)
        if idx is not None:
            up = up_data[idx]
            up['testing_rev_total'] = total
            up['testing_rev_yearly'] = dict(sorted(up_yearly_rev[up_name].items()))
            up['testing_rev_sparkline'] = [round(up_monthly_rev[up_name].get(m, 0), 2) for m in sparkline_months]
            up['testing_rev_months'] = sparkline_months
            lytd, tytd, nrr = compute_ytd_metrics(up_monthly_rev[up_name])
            up['testing_rev_lytd'] = lytd
            up['testing_rev_tytd'] = tytd
            up['testing_rev_nrr'] = nrr
            acc_ids = up_groups.get(up['id'], [])
            up['testing_rev_target_2026'] = sum(
                (accounts[aid].get('rev_target_2026', 0) or 0) for aid in acc_ids
            )
            matched += 1

    # Enrich hierarchy accounts with testing revenue
    for up in up_data:
        for acc in up.get('hierarchy', []):
            acc_name = acc['name']
            if acc_name in acc_total_rev:
                acc['testing_rev_total'] = round(acc_total_rev[acc_name], 2)
                acc['testing_rev_yearly'] = dict(sorted(acc_yearly_rev[acc_name].items()))
                acc['testing_rev_sparkline'] = [round(acc_monthly_rev[acc_name].get(m, 0), 2) for m in sparkline_months]
                lytd, tytd, nrr = compute_ytd_metrics(acc_monthly_rev[acc_name])
                acc['testing_rev_lytd'] = lytd
                acc['testing_rev_tytd'] = tytd
                acc['testing_rev_nrr'] = nrr

    # Testing-only UPs (have testing revenue but no contract items)
    testing_only_ups = []
    for up_name, total in sorted(up_total_rev.items(), key=lambda x: -x[1]):
        if up_name not in up_name_to_idx:
            lytd, tytd, nrr = compute_ytd_metrics(up_monthly_rev[up_name])
            testing_only_ups.append({
                'name': up_name,
                'testing_rev_total': total,
                'testing_rev_yearly': dict(sorted(up_yearly_rev[up_name].items())),
                'testing_rev_sparkline': [round(up_monthly_rev[up_name].get(m, 0), 2) for m in sparkline_months],
                'testing_rev_months': sparkline_months,
                'testing_rev_lytd': lytd,
                'testing_rev_tytd': tytd,
                'testing_rev_nrr': nrr,
            })

    print(f"    {matched} matched (ARR+testing), {len(testing_only_ups)} testing-only")
    return up_data, testing_only_ups, sparkline_months


# ═══════════════════════════════════════════════════════════════
#  BCL (Revenue Recon tab) — built from CI report + samples
# ═══════════════════════════════════════════════════════════════

def extract_big_customer_list(accounts, all_ci, samples_data, acc_casesafe_to_up):
    """Build BCL-equivalent data from CIs and samples.

    Groups CIs by UP, computes B1/B2 ARR, enriches with testing revenue metrics.
    Produces the same JSON structure as the old BCL sheet.
    """
    print("  Building big customer list...")
    today = datetime.date.today()
    current_year = today.year
    current_month = today.month

    company_monthly_rev = samples_data['company_monthly_rev']
    up_monthly_rev = samples_data['up_monthly_rev']
    up_yearly_rev = samples_data['up_yearly_rev']
    up_total_rev = samples_data['up_total_rev']

    # Group CIs by UP ID
    up_ci = defaultdict(list)
    for ci in all_ci:
        if ci['up_id']:
            up_ci[ci['up_id']].append(ci)

    # Group accounts by UP ID
    up_groups = defaultdict(list)
    for acc_id, info in accounts.items():
        if info['up_id']:
            up_groups[info['up_id']].append(acc_id)

    # Month headers (last 27 months for the chart)
    month_headers = []
    for i in range(26, -1, -1):
        m = current_month - i
        y = current_year
        while m <= 0:
            m += 12
            y -= 1
        month_headers.append(f"{y}-{m:02d}")

    # Build one row per UP
    bcl_rows = []
    seen_ups = set()

    for up_id, cis in up_ci.items():
        if not up_id or up_id in seen_ups:
            continue
        seen_ups.add(up_id)

        up_info = accounts.get(up_id, {})
        up_name = up_info.get('name', '')
        if not up_name:
            up_name = cis[0].get('up_name', up_id) if cis else up_id

        # B1/B2 ARR from CIs
        b1_arr = sum(ci['arr_b1'] for ci in cis if ci['bucket'] == 'B1')
        b2_arr = sum(ci['arr_b2'] for ci in cis if ci['bucket'] == 'B2')
        total_arr = b1_arr + b2_arr

        # Account-level info — start from UP record itself, then fill from children
        acc_ids = up_groups.get(up_id, [])
        csm = up_info.get('csm', '') or ''
        owner = up_info.get('owner', '') or ''
        industry = up_info.get('industry', '') or ''
        tam_type = up_info.get('tam_type', '') or ''
        account_cohort = up_info.get('account_cohort', '') or ''
        for aid in acc_ids:
            a = accounts.get(aid, {})
            if a.get('csm') and not csm:
                csm = a['csm']
            if a.get('owner') and not owner:
                owner = a['owner']
            if a.get('industry') and not industry:
                industry = a['industry']
            if a.get('tam_type') and not tam_type:
                tam_type = a['tam_type']
            if a.get('account_cohort') and not account_cohort:
                account_cohort = a['account_cohort']

        # Monthly testing revenue
        monthly = [round(up_monthly_rev.get(up_name, {}).get(m, 0), 2) for m in month_headers]

        # Testing metrics
        test_data = up_monthly_rev.get(up_name, {})
        lytd = sum(test_data.get(f"{current_year - 1}-{mo:02d}", 0) for mo in range(1, current_month + 1))
        tytd = sum(test_data.get(f"{current_year}-{mo:02d}", 0) for mo in range(1, current_month + 1))
        nrr = round((tytd / lytd - 1) * 100, 1) if lytd > 0 else 0

        l12m = sum(test_data.get(m, 0) for m in month_headers[-12:])
        l6m = sum(test_data.get(m, 0) for m in month_headers[-6:])

        fy24 = sum(test_data.get(f"2024-{mo:02d}", 0) for mo in range(1, 13))
        fy25 = sum(test_data.get(f"2025-{mo:02d}", 0) for mo in range(1, 13))

        # Activity metrics
        active_months_list = [m for m in month_headers if test_data.get(m, 0) > 0]
        active_months = len(active_months_list)
        frequency = round(active_months / len(month_headers), 2) if month_headers else 0

        # First/last test dates
        all_months_with_data = sorted(k for k, v in test_data.items() if v > 0)
        first_test = all_months_with_data[0] + '-01' if all_months_with_data else ''
        last_test = all_months_with_data[-1] + '-01' if all_months_with_data else ''

        # Months since first/last
        if first_test:
            ft_parts = first_test.split('-')
            months_since_first = (current_year - int(ft_parts[0])) * 12 + (current_month - int(ft_parts[1]))
        else:
            months_since_first = 0
        if last_test:
            lt_parts = last_test.split('-')
            months_since_last = (current_year - int(lt_parts[0])) * 12 + (current_month - int(lt_parts[1]))
        else:
            months_since_last = 0

        # Simple trend calculation (slope of last N months)
        def trend(n):
            vals = [test_data.get(m, 0) for m in month_headers[-n:]]
            if not any(vals):
                return ''
            ups = sum(1 for i in range(1, len(vals)) if vals[i] > vals[i-1])
            downs = sum(1 for i in range(1, len(vals)) if vals[i] < vals[i-1])
            if ups > downs:
                return 'Up'
            elif downs > ups:
                return 'Down'
            return 'Flat'

        # L12M / L6M averages
        l12m_avg = l12m / 12 if l12m else 0
        l6m_avg = l6m / 6 if l6m else 0
        ratio_12v6 = round(l6m_avg / l12m_avg, 2) if l12m_avg > 0 else 0

        # Velocity & momentum
        velocity = round(tytd / (current_month or 1), 2)
        momentum = round(l6m_avg - l12m_avg, 2) if l12m_avg else 0

        # Score (simple: weighted combo of tytd + trend)
        score_rank = round(tytd / 1000, 1) if tytd else 0

        # Status
        if total_arr > 0 and tytd > 0:
            status = 'Active'
        elif total_arr > 0:
            status = 'ARR Only'
        elif tytd > 0:
            status = 'Testing Only'
        elif months_since_last > 6:
            status = 'Churned'
        else:
            status = 'Inactive'

        # Growth cohort
        if nrr > 20:
            growth_cohort = 'Growing'
        elif nrr < -20:
            growth_cohort = 'Declining'
        elif nrr != 0:
            growth_cohort = 'Stable'
        else:
            growth_cohort = 'New/Unknown'

        # Tenure: New Logo = first test within last 12 months, else Established Business
        tenure = 'New Logo' if months_since_first < 12 else 'Established Business'

        # Performance quadrant
        if total_arr > 50000 and tytd > 20000:
            perf_quad = 'Stars'
        elif total_arr > 50000:
            perf_quad = 'Cash Cows'
        elif tytd > 20000:
            perf_quad = 'Rising'
        else:
            perf_quad = 'Watch'

        # Credit balance
        credit_bal = sum(
            (accounts.get(aid, {}).get('credit_balance', 0) or 0) for aid in acc_ids
        )

        # Salesforce URL
        sf_url = (
            f'https://cambri.lightning.force.com/lightning/r/Account/{up_id}/view'
            if up_id and len(up_id) > 10
            else ''
        )

        rec = {
            'csm': csm,
            'owner': owner,
            'cls': tam_type or '',
            'industry': industry,
            'up': up_name,
            'monthly': monthly,
            'status': status,
            'l12m': round(l12m, 2),
            'l6m': round(l6m, 2),
            'lytd': round(lytd, 2),
            'tytd': round(tytd, 2),
            'nrr': round(nrr, 1),
            'fy24': round(fy24, 2),
            'fy25': round(fy25, 2),
            'fc25': 0,
            'target26': 0,
            'perf_quad': perf_quad,
            'rev_gap': 0,
            'ly_vs_ty': round((tytd / lytd - 1) * 100, 1) if lytd > 0 else (100.0 if tytd > 0 else 0),
            'ytd_vs_tgt': 0,
            'growth_cohort': growth_cohort,
            'account_cohort': account_cohort,
            'tenure': tenure,
            'trend_18m': trend(18),
            'trend_12m': trend(12),
            'trend_6m': trend(6),
            'activity': 'Active' if months_since_last <= 2 else ('Recent' if months_since_last <= 6 else 'Dormant'),
            'active_months': active_months,
            'frequency': frequency,
            'first_test': first_test,
            'months_since_first': months_since_first,
            'last_test': last_test,
            'months_since_last': months_since_last,
            'velocity': velocity,
            'h1_24': round(sum(test_data.get(f"2024-{mo:02d}", 0) for mo in range(1, 7)), 2),
            'l12m_avg': round(l12m_avg, 2),
            'l6m_avg': round(l6m_avg, 2),
            'ratio_12v6': ratio_12v6,
            'momentum': momentum,
            'score_rank': score_rank,
            'lic_fy24': 0,
            'lic_ytd': 0,
            'cred_fy24': 0,
            'cred_ytd': 0,
            'ms_fy24': 0,
            'ms_ytd': 0,
            'test_fy24': round(fy24, 2),
            'test_ytd': round(tytd, 2),
            'total_fy24': round(fy24, 2),
            'total_fy25': round(fy25, 2),
            'arr_calc': round(total_arr, 2),
            'hierarchy_live_arr': round(up_info.get('hierarchy_total_arr', 0) or 0, 2),
            'eoq4_arr': round(up_info.get('eoq4_total', 0) or 0, 2),
            'pending': 0,
            'up_id': up_id,
            'tam_type': tam_type,
            'credit_bal': round(credit_bal, 2),
            'sf_url': sf_url,
        }
        bcl_rows.append(rec)

    # Also add UPs that have testing revenue but no CIs
    for up_name, total in sorted(up_total_rev.items(), key=lambda x: -x[1]):
        # Find UP ID from accounts
        up_id_found = ''
        for acc_id, info in accounts.items():
            if info['name'] == up_name and info['up_id']:
                up_id_found = info['up_id']
                break
        if up_id_found and up_id_found in seen_ups:
            continue
        if up_id_found:
            seen_ups.add(up_id_found)

        test_data = up_monthly_rev.get(up_name, {})
        monthly = [round(test_data.get(m, 0), 2) for m in month_headers]
        lytd = sum(test_data.get(f"{current_year - 1}-{mo:02d}", 0) for mo in range(1, current_month + 1))
        tytd = sum(test_data.get(f"{current_year}-{mo:02d}", 0) for mo in range(1, current_month + 1))
        nrr = round((tytd / lytd - 1) * 100, 1) if lytd > 0 else 0

        sf_url = (
            f'https://cambri.lightning.force.com/lightning/r/Account/{up_id_found}/view'
            if up_id_found and len(up_id_found) > 10
            else ''
        )

        rec = {
            'csm': '', 'owner': '', 'cls': '', 'industry': '',
            'up': up_name,
            'monthly': monthly,
            'status': 'Testing Only',
            'l12m': round(sum(test_data.get(m, 0) for m in month_headers[-12:]), 2),
            'l6m': round(sum(test_data.get(m, 0) for m in month_headers[-6:]), 2),
            'lytd': round(lytd, 2),
            'tytd': round(tytd, 2),
            'nrr': round(nrr, 1),
            'fy24': round(sum(test_data.get(f"2024-{mo:02d}", 0) for mo in range(1, 13)), 2),
            'fy25': round(sum(test_data.get(f"2025-{mo:02d}", 0) for mo in range(1, 13)), 2),
            'fc25': 0, 'target26': 0, 'perf_quad': 'Watch', 'rev_gap': 0,
            'ly_vs_ty': round((tytd / lytd - 1) * 100, 1) if lytd > 0 else (100.0 if tytd > 0 else 0), 'ytd_vs_tgt': 0,
            'growth_cohort': 'New/Unknown', 'account_cohort': '', 'tenure': 'New Logo',
            'trend_18m': '', 'trend_12m': '', 'trend_6m': '',
            'activity': '', 'active_months': 0, 'frequency': 0,
            'first_test': '', 'months_since_first': 0,
            'last_test': '', 'months_since_last': 0,
            'velocity': 0, 'h1_24': 0, 'l12m_avg': 0, 'l6m_avg': 0,
            'ratio_12v6': 0, 'momentum': 0, 'score_rank': 0,
            'lic_fy24': 0, 'lic_ytd': 0, 'cred_fy24': 0, 'cred_ytd': 0,
            'ms_fy24': 0, 'ms_ytd': 0, 'test_fy24': 0, 'test_ytd': round(tytd, 2),
            'total_fy24': 0, 'total_fy25': 0, 'arr_calc': 0,
            'hierarchy_live_arr': 0, 'eoq4_arr': 0,
            'pending': 0, 'up_id': up_id_found, 'tam_type': '', 'credit_bal': 0,
            'sf_url': sf_url,
        }
        bcl_rows.append(rec)

    # Sort by total ARR descending, then by tytd
    bcl_rows.sort(key=lambda x: -(x['arr_calc'] + x['tytd']))

    print(f"    {len(bcl_rows)} BCL rows, {len(month_headers)} months")

    return {
        'month_headers': month_headers,
        'rows': bcl_rows,
        'last_updated': today.strftime('%Y-%m-%d %H:%M'),
    }


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    """Main entry point."""
    print("Starting compute job...")
    today = datetime.date.today()

    try:
        # Download workbook
        print("\nDownloading ARR working file...")
        wb = download_workbook(SOURCE_URL, 'ARR Working File')

        # 1. Read accounts
        print("\nReading accounts...")
        accounts, id15_to_id18, acc_casesafe_to_up = read_accounts(wb)

        # 2. Read contract items
        print("\nReading contract items...")
        all_ci = read_contract_items(wb, acc_casesafe_to_up)

        # 3. Read samples
        print("\nReading samples...")
        samples_data = read_samples(wb, acc_casesafe_to_up, accounts)

        wb.close()

        # 4. Build CEO dashboard
        print("\nBuilding outputs...")
        ceo_data = extract_ceo_dashboard(samples_data, accounts, acc_casesafe_to_up)

        # 5. Build UP Explorer
        up_data, testing_only_ups, sparkline_months = extract_up_explorer(
            accounts, id15_to_id18, acc_casesafe_to_up, all_ci, samples_data
        )

        # 6. Build BCL
        bcl_data = extract_big_customer_list(accounts, all_ci, samples_data, acc_casesafe_to_up)

        # Build output
        print("\nBuilding output JSON...")
        full_output = {
            'up_data': up_data,
            'testing_only_ups': testing_only_ups,
            'sparkline_months': sparkline_months,
            'bcl': bcl_data,
            'ceo': ceo_data,
            'meta': {
                'extracted': today.isoformat(),
                'source': 'ARR Working File',
                'total_ups_arr': len(up_data),
                'testing_only_count': len(testing_only_ups),
            },
        }

        # Write JSON
        output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dashboard_data.json')
        output_json = json.dumps(full_output, indent=2, ensure_ascii=False)
        with open(output_path, 'w') as f:
            f.write(output_json)

        print(f"\nSuccess!")
        print(f"  Wrote {len(output_json)} bytes to {output_path}")
        print(f"  UP Explorer: {len(up_data)} UPs, {len(testing_only_ups)} testing-only")
        print(f"  BCL: {len(bcl_data['rows'])} customers, {len(bcl_data['month_headers'])} months")
        print(f"  CEO: {len(ceo_data['months'])} months")

    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == '__main__':
    exit(main())
